import time
import os
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys


# Setting up Chrome Options
chrome_options = Options()
# chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--start-maximized')
# chrome_options.add_argument('--single-process')
# chrome_options.add_argument('--disable-dev-shm-usage')
# chrome_options.add_argument("--incognito")
# chrome_options.add_argument('--disable-blink-features=AutomationControlled')
chrome_options.add_argument('--disable-blink-features=AutomationControlled')
# chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
# chrome_options.add_argument("disable-infobars")
# chrome_options.add_experimental_option( "prefs",{'profile.managed_default_content_settings.javascript': 2})
chrome_options.add_experimental_option("detach", True)
prefs = {"profile.default_content_setting_values.notifications": 2}
chrome_options.add_experimental_option("prefs", prefs)
# New Update in Selenium 4. Refer to documentation
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()),options=chrome_options)

# Opening the required excel files
cwd = os.getcwd() + "\\Excels\\"
book1 = openpyxl.load_workbook(cwd + "2. MASTER SHEET.xlsx")
book2 = openpyxl.load_workbook(cwd + "1. Account Credentials.xlsx")
returnedBook = openpyxl.load_workbook(cwd + "3. Returned Orders.xlsx")
cancelledBook = openpyxl.load_workbook(cwd + "4. Cancelled.xlsx")
miscBook = openpyxl.load_workbook(cwd + "5. Miscellaneous.xlsx")
statusNotChanged = openpyxl.load_workbook(cwd + "6. Status Not Changed.xlsx")

# Opening Lazada Website
def openLazada():
    driver.get("https://sellercenter.lazada.com.my/apps/seller/login")

# Providing the Login Credentials
def account_login(userID,password):
    print("Inside Account Login " + userID + " " + password)
    time.sleep(7)
    # Providing User ID
    driver.find_element(By.ID, "account").send_keys(Keys.CONTROL + "a")
    driver.find_element(By.ID, "account").send_keys(Keys.DELETE)
    driver.find_element(By.ID, "account").send_keys(userID)
    # Providing Password
    driver.find_element(By.ID, "password").send_keys(Keys.CONTROL + "a")
    driver.find_element(By.ID, "password").send_keys(Keys.DELETE)
    driver.find_element(By.ID,"password").send_keys(password)
    driver.find_element(By.XPATH,"//span[text()='Login']").click()
    time.sleep(10)
    # Checking for Unsuccessful Login
    try:
        driver.find_element(By.XPATH,"//*[@id='J_page']/div[2]/div/div[2]/div/div[2]/div/form/div[1]/div/div")
    except NoSuchElementException:
        # Login Successful
        return 1
    return 0

def open_Order_section():
    print("Inside Order Section")
    time.sleep(5)
    # Close any ads Pop-Up
    # driver.find_element(By.TAG_NAME,"body").send_keys(Keys.ESCAPE)
    # try:
    #     driver.find_element(By.XPATH,"//div[@class='next-overlay-backdrop next-dialog-container']/div/a")
    # except NoSuchElementException:
    #     print("No Pop-Ups found")
    # Click On Orders Dropdown
    try:
        driver.find_element(By.XPATH, "//*[@id='layout-new-menu-content']/li[2]/div/div/i").click()
    except NoSuchElementException:
        # Due to smaller screen size clicking in the lazada image
        driver.find_element(By.XPATH, "/html/body/div[1]/section/div[1]/aside/div[1]/a").click()

    # Click On Orders from dropdown
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, "//*[@id='layout-new-menu-content']/li[2]/ul/a[1]").click()
    except NoSuchElementException:
    #     Small Screen issue: After clicking on logo.. clicking on pending order
        driver.find_element(By.XPATH, "//div[text()='Pending Orders']").click()
    time.sleep(2)

    # Click on All section
    driver.find_element(By.XPATH, "//*[@id='root']/section/div[2]/div/div[1]/div/div/div[3]/div/div[1]/div/div/div/ul/li[1]/div/span").click()

# Searching for the Order Status
def order_status(track_id):
    print("Inside Order Status")
    time.sleep(2)
    if track_id == None:
        status = "INCORRECT TN NUMBER! Please Check Again!"
        orderID = "INCORRECT TN NUMBER! Please Check Again!"
        return status, orderID
    driver.find_element(By.XPATH,"//input[@name='trackingNumber']").send_keys(Keys.BACKSPACE)
    time.sleep(1)
    driver.find_element(By.XPATH,"//input[@name='trackingNumber']").send_keys(track_id)
    driver.find_element(By.XPATH, "//input[@name='trackingNumber']").send_keys(Keys.ENTER)
    time.sleep(4)
    try:
        orderID = driver.find_element(By.XPATH,"//*[@id='root']/section/div[2]/div/div[1]/div/div/div[4]/div/div[3]/div/div[1]/div[2]/div[2]/span[1]/span[2]/span/a").text
        status = driver.find_element(By.XPATH,"//*[@id='root']/section/div[2]/div/div[1]/div/div/div[4]/div/div[3]/div/div[2]/div/div/div[4]/div[1]/span[2]").text
        print(status)
        print(orderID)
        print()
    except NoSuchElementException:
        status = "INCORRECT TN NUMBER! Please Check Again!"
        orderID = "INCORRECT TN NUMBER! Please Check Again!"
    return status, orderID

# Log Out OF the Account
def log_out():
    print("Inside Log Out")
    driver.find_element(By.XPATH,"//*[@id='navi_right_sidebar_id']/div[2]/div/div[3]/span[1]/i").click()
    driver.find_element(By.XPATH,"//*[@id='navi_right_sidebar_id']/div[2]/div[2]/ul/li[2]/div/span/span/button/span").click()
    return 1

# Acquire all the available sheets in "MASTER SHEET" File
def get_sheet_name():
    sheetNames = book1.sheetnames
    return sheetNames

# Aquire the Credentials of the specified Account
def get_account_details(accountName):
    sheet = book2.active
    for i in range(2,sheet.max_row+1):
        if sheet.cell(row=i, column=2).value == accountName:
            userID = sheet.cell(row=i, column=3).value
            password = sheet.cell(row=i, column=4).value
            return userID, password
    return 0,0

# Update the status in the excel sheet by calling the functions
def update_books(accountName):
    # print("Inside Udate Status")
    # print(accountName)
    book1.active = book1[accountName]
    sheet1 = book1.active
    sheet2 = book2.active
    userID,password = get_account_details(accountName)
    if userID == 0:
        return
    result = account_login(userID,password)

    # If Account login is Unsuccessful print it in Accounts sheet
    if result == 0:
        # print("Login Unsuccessful for the account " + accountName)
        for i in range(2,sheet2.max_row+1):
            if sheet2.cell(row=i, column=2).value == accountName:
                sheet2.cell(row=i, column=5).value = "WRONG CREDENTIALS! Please Try Again"
                book2.save(cwd + "1. Account Credentials.xlsx")
                return
    elif result == 1:
        for i in range(2,sheet2.max_row+1):
            if sheet2.cell(row=i, column=2).value == accountName:
                sheet2.cell(row=i, column=5).value = "Account Login Successful"
                book2.save(cwd + "1. Account Credentials.xlsx")

    # Get to the Order Section after successful login
    open_Order_section()
    for i in range(2,sheet1.max_row+1):
        # Getting current date and time
        now = datetime.now()

        tn = sheet1.cell(row=i,column=4).value
        status, orderID = order_status(tn)
        sheet1.cell(row=i, column=6).value = status
        sheet1.cell(row=i,column=10).value = orderID
        sheet1.cell(row=i, column=5).value = now.strftime("%d/%B/%Y - %I:%M %p")
        if status == "Delivered":
            sheet1.cell(row=i, column=7).value = "No Attention Required"
            if sheet1.cell(row=i, column=8).value == "NO":
                sheet1.cell(row=i, column=7).value = "ATTENTION!"
        elif status != "Delivered" and sheet1.cell(row=i, column=8).value != "YES":
            sheet1.cell(row=i, column=7).value = "ATTENTION!"

    log_out()
    book1.save(cwd+"2. MASTER SHEET.xlsx")
    book2.save(cwd+"1. Account Credentials.xlsx")
    # print("saved Successfully")


# Status of shipped is :- "Shipped"
# Append the Return Sheet
def append_filter_sheet(mode,accountName,tn,orderID):
    # Check which sheet to write
    if mode == "Returned":
        returnedBook.active = returnedBook[accountName]
        sheet = returnedBook.active
    elif mode == "Cancelled":
        cancelledBook.active = cancelledBook[accountName]
        sheet = cancelledBook.active
    elif mode == "Misc":
        miscBook.active = miscBook[accountName]
        sheet = miscBook.active
    elif mode == "StatusNotChanged":
        statusNotChanged.active = statusNotChanged[accountName]
        sheet = statusNotChanged.active

    # Getting maximum rows present
    rowNumber = sheet.max_row+1
    # Getting current date and time
    now = datetime.now()
    now_time = now.strftime("%d/%B/%Y - %I:%M %p")

    # Do Nothing if the Order is already present only update time
    for i in range(2,rowNumber):
        if tn == sheet.cell(row=i,column=1).value :
            sheet.cell(row=i, column=2).value = now_time
            return

    # Populate Transaction Number
    sheet.cell(row=rowNumber,column=1).value = tn
    # Populate Updation Date and time
    sheet.cell(row=rowNumber,column=2).value = now_time
    # Populate Remarks
    sheet.cell(row=rowNumber, column=3).value = "ATTENTION!"
    # Populate OrderID
    sheet.cell(row=rowNumber, column=6).value = orderID

# Filter out the returned/Returning Products
def filter_orders(accountName):
    # Opening the required sheets
    book1.active = book1[accountName]
    sheet1 = book1.active

    # Filtering Out from the MASTER SHEET
    for i in range(2,sheet1.max_row+1):
        status = sheet1.cell(row=i, column=6).value

        # Filtering out the Returned Orders
        if status == "returned" or status == "In Transit: Returning to seller"\
                or status == "Package Returned":
            tn = sheet1.cell(row=i,column=4).value
            orderID = sheet1.cell(row=i,column=10).value
            append_filter_sheet("Returned",accountName,tn,orderID)

        # Filtering out the Cancelled Orders
        if status == "Canceled":
            tn = sheet1.cell(row=i,column=4).value
            orderID = sheet1.cell(row=i,column=10).value
            append_filter_sheet("Cancelled",accountName,tn,orderID)

        # Filtering out the Misc Orders
        if status == "Lost by 3PL" or status == "Package scrapped"\
                or status == "Damaged By 3PL" :
            tn = sheet1.cell(row=i, column=4).value
            orderID = sheet1.cell(row=i, column=10).value
            append_filter_sheet("Misc", accountName, tn, orderID)

        # Filtering out the Status Not Changed Orders
        if status == "Pending" or status == "Packed" or status == "Ready To Ship":
            tn = sheet1.cell(row=i, column=4).value
            orderID = sheet1.cell(row=i, column=10).value
            append_filter_sheet("StatusNotChanged", accountName, tn, orderID)

# Function for changing the Remarks of those status which have been resolved
def attention_checker(accountName):
    # Open all the filter books
    returnedBook.active = returnedBook[accountName]
    cancelledBook.active = cancelledBook[accountName]
    miscBook.active = miscBook[accountName]
    statusNotChanged.active = statusNotChanged[accountName]

    sheet1 = returnedBook.active
    sheet2 = cancelledBook.active
    sheet3 = miscBook.active
    sheet4 = statusNotChanged.active
    # iterate in the books and change the Remarks Column
    for i in range(2,sheet1.max_row+1):
        if sheet1.cell(row=i,column=4).value == "YES":
            sheet1.cell(row=i,column=3).value = "Issue Resolved"
    for i in range(2,sheet2.max_row+1):
        if sheet2.cell(row=i,column=4).value == "YES":
            sheet2.cell(row=i,column=3).value = "Issue Resolved"
    for i in range(2,sheet3.max_row+1):
        if sheet3.cell(row=i,column=4).value == "YES":
            sheet3.cell(row=i,column=3).value = "Issue Resolved"
    for i in range(2,sheet4.max_row+1):
        if sheet4.cell(row=i,column=4).value == "YES":
            sheet4.cell(row=i,column=3).value = "Issue Resolved"


# Function for updating the Run Time insdie the Log Book
def update_logbook():
    logbook = openpyxl.load_workbook(cwd + "Log Book.xlsx")
    sheet = logbook.active
    # Getting current date and time
    now = datetime.now()
    now_time = now.strftime("%d/%B/%Y - %I:%M %p")
    sheet.cell(row=sheet.max_row+1,column=1).value = now_time
    logbook.save(cwd + "Log Book.xlsx")

# Performing The Steps in the Main

# Step-1: Opening Lazada
openLazada()

# Step-2: Getting Account names
accounts = get_sheet_name()

# step-3: Updating the Status of all the Available Accounts in the "MASTER SHEET" file
for i in accounts:
    update_books(i)

# step-4: Filtering out the Returned Orders
for i in accounts:
    print(i);
    # filter_orders(i)
    # attention_checker(i)

# Updating the Run Time in Log Book
update_logbook()

# Step-4: Saving the Excel Files
book1.save(cwd + "2. MASTER SHEET.xlsx")
book2.save(cwd + "1. Account Credentials.xlsx")
returnedBook.save(cwd + "3. Returned Orders.xlsx")
cancelledBook.save(cwd + "4. Cancelled.xlsx")
miscBook.save(cwd + "5. Miscellaneous.xlsx")
statusNotChanged.save(cwd + "6. Status Not Changed.xlsx")

# Closing the Books and the browser
book1.close()
book2.close()
returnedBook.close()
cancelledBook.close()
miscBook.close()
driver.quit()
print("********************************************************************************")
print("*                                 SUCCESSFULL                                  *")
print("********************************************************************************")
