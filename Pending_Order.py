import os
import time
from datetime import datetime
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys

# Setting up Chrome Options
chrome_options = Options()
# chrome_options.add_argument('--no-sandbox')
chrome_options.add_argument('--start-maximized')
# chrome_options.add_argument('--single-process')
# chrome_options.add_argument('--disable-dev-shm-usage')
# chrome_options.add_argument("--incognito")
# chrome_options.add_argument('--disable-blink-features=AutomationControlled')
chrome_options.add_argument('--disable-blink-features=AutomationControlled')
# chrome_options.add_experimental_option('useAutomationExtension', False)
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
# chrome_options.add_argument("disable-infobars")
# chrome_options.add_experimental_option( "prefs",{'profile.managed_default_content_settings.javascript': 2})
chrome_options.add_experimental_option("detach", True)
prefs = {"profile.default_content_setting_values.notifications": 2}
chrome_options.add_experimental_option("prefs", prefs)

# New Update in Selenium 4. Refer to documentation
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)

# getting the object of today
today = datetime.today()
# Conversion of today object to string
today_date = today.strftime("%d %b %Y %H:%M")

# Opening the required excel files
cwd = os.getcwd() + "\\Excels\\"
orderBook = openpyxl.load_workbook(cwd + "Pending Orders.xlsx")
accountBook = openpyxl.load_workbook(cwd + "Account Credentials.xlsx")


# Opening Lazada Website
def openLazada():
    driver.get("https://sellercenter.lazada.com.my/apps/seller/login")


# Aquire the Credentials of the specified Account
def get_account_details(accountName):
    sheet = accountBook.active
    for i in range(2, sheet.max_row + 1):
        if sheet.cell(row=i, column=2).value == accountName:
            userID = sheet.cell(row=i, column=3).value
            password = sheet.cell(row=i, column=4).value
            return userID, password
    return 0, 0


# Providing the Login Credentials
def account_login(userID, password):
    time.sleep(7)
    driver.find_element(By.ID, "account").send_keys(Keys.CONTROL + "a")
    driver.find_element(By.ID, "account").send_keys(Keys.DELETE)
    driver.find_element(By.ID, "account").send_keys(userID)
    driver.find_element(By.ID, "password").send_keys(Keys.CONTROL + "a")
    driver.find_element(By.ID, "password").send_keys(Keys.DELETE)
    driver.find_element(By.ID, "password").send_keys(password)
    driver.find_element(By.XPATH, "//span[text()='Login']").click()
    time.sleep(10)
    try:
        driver.find_element(By.XPATH, "//*[@id='J_page']/div[2]/div/div[2]/div/div[2]/div/form/div[1]/div/div")
    except NoSuchElementException:
        # Login Successful
        return 1
    return 0


# Opening the ToShip Order Section
def open_Order_section():
    time.sleep(5)
    # Close any ads Pop-Up
    # driver.find_element(By.TAG_NAME,"body").send_keys(Keys.ESCAPE)
    # try:
    #     driver.find_element(By.XPATH,"//div[@class='next-overlay-backdrop next-dialog-container']/div/a")
    # except NoSuchElementException:
    #     print("No Pop-Ups found")
    # Click On Orders Dropdown
    try:
        driver.find_element(By.XPATH, "//*[@id='layout-new-menu-content']/li[2]/div/div/i").click()
    except NoSuchElementException:
        # Due to smaller screen size clicking in the lazada image
        driver.find_element(By.XPATH, "/html/body/div[1]/section/div[1]/aside/div[1]/a").click()

    # Click On Orders from dropdown
    time.sleep(2)
    try:
        driver.find_element(By.XPATH, "//*[@id='layout-new-menu-content']/li[2]/ul/a[1]").click()
    except NoSuchElementException:
    #     Small Screen issue: After clicking on logo.. clicking on pending order
        driver.find_element(By.XPATH, "//div[text()='Pending Orders']").click()
    time.sleep(2)


# Log Out OF the Account
def log_out():
    driver.find_element(By.XPATH, "//*[@id='navi_right_sidebar_id']/div[2]/div/div[3]/span[1]/i").click()
    driver.find_element(By.XPATH,
                        "//*[@id='navi_right_sidebar_id']/div[2]/div[2]/ul/li[2]/div/span/span/button/span").click()
    # driver.find_element(By.TAG_NAME,'body').send_keys(Keys.CONTROL + 't')
    return 1


# Difference between days
def days_elapsed(time):
    # Conversion of string to datetime obj
    time_obj = datetime.strptime(time, "%d %b %Y %H:%M")
    # getting the object of today
    today = datetime.today()
    # Difference between days of two objects
    days = today - time_obj
    # Result of difference
    return days.days


# Updating details from ToPack Section (For 20 Orders)
def update_ToPack(accountName):
    time.sleep(5)
    path = "tab_topack_order_list_"
    time.sleep(2)
    # Selecting the arrow to change Order number per page
    driver.find_element(By.XPATH,"//*[@id='root']/section/div[2]/div/div[1]/div/div/div[4]/div/div[4]/div[2]/span[2]/span[1]/span[2]/span/i").click()
    # Select number 80
    time.sleep(3)
    driver.find_element(By.XPATH, "//span[text()='80']").click()
    # /html/body/div[3]/div/ul/li[4]
    time.sleep(3)
    for i in range(0, 81):
        path1 = "'" + path + str(i) + "'"
        xpath1 = "//div[@data-spm=" + path1 + "]/div[1]/div[2]/div[2]/span[1]/span[2]"
        try:
            orderID = driver.find_element(By.XPATH, xpath1).text
        except NoSuchElementException:
            return
        xpath2 = "//div[@data-spm=" + path1 + "]/div[1]/div[2]/div[2]/span[2]/span[2]"
        time.sleep(1)
        creation_time = driver.find_element(By.XPATH, xpath2).text
        append_pending_sheet(accountName, "To Pack", orderID, creation_time, days_elapsed(creation_time))
        print(orderID)
        print(creation_time)
        print("days passed ",days_elapsed(creation_time) )


# Updating details from ToArrangeShipment Section
def update_ToArrange(accountName):
    time.sleep(5)
    # print("Inside ToArrange")
    driver.find_element(By.XPATH,
                        "//*[@id='root']/section/div[2]/div/div[1]/div/div/form/div[1]/div/div/label[2]").click()
    path = "tab_toshiparrangeshipment_order_list_"
    time.sleep(2)
    for i in range(0, 81):
        path1 = "'" + path + str(i) + "'"
        xpath1 = "//div[@data-spm=" + path1 + "]/div[1]/div[2]/div[2]/span[1]/span[2]"
        try:
            orderID = driver.find_element(By.XPATH, xpath1).text
        except NoSuchElementException:
            # print("No Orders found.hitting break")
            return
        xpath2 = "//div[@data-spm=" + path1 + "]/div[1]/div[2]/div[2]/span[2]/span[2]"
        time.sleep(1)
        creation_time = driver.find_element(By.XPATH, xpath2).text
        append_pending_sheet(accountName, "To Arrange", orderID, creation_time, days_elapsed(creation_time))
        # print(orderID)
        # print(creation_time)
        # print("days passed ", days_elapsed(creation_time))


# Updating details from ToHandover Section
def update_Tohandover(accountName):
    time.sleep(5)
    # print("Inside ToHandover")
    driver.find_element(By.XPATH,
                        "//*[@id='root']/section/div[2]/div/div[1]/div/div/form/div[1]/div/div/label[3]/span[2]/div").click()
    path = "tab_toshiphandover_order_list_"
    time.sleep(2)
    for i in range(0, 81):
        path1 = "'" + path + str(i) + "'"
        xpath1 = "//div[@data-spm=" + path1 + "]/div[1]/div[2]/div[2]/span[1]/span[2]"
        try:
            orderID = driver.find_element(By.XPATH, xpath1).text
        except NoSuchElementException:
            # print("No Orders found.hitting break")
            return
        xpath2 = "//div[@data-spm=" + path1 + "]/div[1]/div[2]/div[2]/span[2]/span[2]"
        time.sleep(1)
        creation_time = driver.find_element(By.XPATH, xpath2).text
        append_pending_sheet(accountName, "To Handover", orderID, creation_time, days_elapsed(creation_time))
        # print(orderID)
        # print(creation_time)
        # print("days passed ", days_elapsed(creation_time))


def clean_pending_book(accountName):
    orderBook.active = orderBook[accountName]
    sheet = orderBook.active
    # Clear the sheet from row-2 to max row
    sheet.delete_rows(2, sheet.max_row+1)
    orderBook.save(cwd + "Pending Orders.xlsx")


# Acquire all the available sheets in "Pending Order" File
def get_sheet_name():
    sheetNames = orderBook.sheetnames
    return sheetNames


# Append the Pending Order Sheet
def append_pending_sheet(accountName, status, orderID, creationTime, daysElapsed):
    # Open the pending Sheet
    orderBook.active = orderBook[accountName]
    sheet = orderBook.active

    # Getting maximum rows present
    rowNumber = sheet.max_row + 1

    # Populate Order ID
    sheet.cell(row=rowNumber, column=1).value = orderID
    # Populate Status
    sheet.cell(row=rowNumber, column=2).value = status
    # Populate Creation Time
    sheet.cell(row=rowNumber, column=3).value = creationTime
    # Populate days Elapsed
    sheet.cell(row=rowNumber, column=4).value = daysElapsed
    # Decide Action
    if daysElapsed >= 2:
        sheet.cell(row=rowNumber, column=5).value = "IMMEDIATE ATTENTION!!!"
    elif daysElapsed >= 1:
        sheet.cell(row=rowNumber, column=5).value = "Keep The Order Ready"
    # Update the time of Updation
    sheet.cell(row=rowNumber, column=6).value = today_date
    # Saving the workbook
    orderBook.save(cwd + "Pending Orders.xlsx")

# Function for updating the Run Time insdie the Log Book
def update_logbook():
    logbook = openpyxl.load_workbook(cwd + "Log Book.xlsx")
    sheet = logbook.active
    # Getting current date and time
    now = datetime.now()
    now_time = now.strftime("%d/%B/%Y - %I:%M %p")
    sheet.cell(row=sheet.max_row+1,column=1).value = now_time
    logbook.save(cwd + "Log Book.xlsx")

# Performing The Steps in the Main


# Step-1: Opening Lazada
openLazada()

# Step-2: Getting Account names
accounts = get_sheet_name()

# Step-3: Iterating the accounts and fetching the orders status
for i in accounts:
    # Clean the sheet before populating
    clean_pending_book(i)
    # Fetching the Account Details
    sheet2 = accountBook.active
    userID, password = get_account_details(i)
    if userID == 0:
        # Get out from the for loop if USERID, PASSWORD doesnt exist in the excel file
        break
    # Account Login
    result = account_login(userID, password)
    # If Account login is Unsuccessful print it in Accounts sheet
    if result == 0:
        for j in range(2, sheet2.max_row + 1):
            if sheet2.cell(row=j, column=2).value == i:
                sheet2.cell(row=j, column=5).value = "WRONG CREDENTIALS! Please Try Again"
                accountBook.save(cwd + "Account Credentials.xlsx")
        continue
    elif result == 1:
        for j in range(2, sheet2.max_row + 1):
            if sheet2.cell(row=j, column=2).value == i:
                sheet2.cell(row=j, column=5).value = "Account Login Successful"
                accountBook.save(cwd + "Account Credentials.xlsx")
    # Opening Order Section
    open_Order_section()
    # Calling Each Section Function
    update_ToPack(i)
    update_ToArrange(i)
    update_Tohandover(i)
    # Log Out
    log_out()
# Close the Browser
driver.quit()
update_logbook()
# openLazada()
# account_login("jbbowlover@gmail.com","@Sachinpanwar83@")
# open_Order_section()
# update_ToPack("jbbowlover@gmail.com")
# update_ToArrange("jbbowlover@gmail.com")
# update_Tohandover("jbbowlover@gmail.com")
# log_out()
# driver.quit()
# accountNames = get_sheet_name()
# for i in accountNames:
#     clean_pending_book(i)

orderBook.save(cwd + "Pending Orders.xlsx")
accountBook.save(cwd + "Account Credentials.xlsx")

print("********************************************************************************")
print("*                                 SUCCESSFULL                                  *")
print("********************************************************************************")
